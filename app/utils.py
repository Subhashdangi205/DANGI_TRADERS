# app/utils.py
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generate_customer_excel(customer, transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = "Khata Report"

    # Header Styling
    header_fill = PatternFill(
        start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # Title Block
    ws.append(["DANGI TRADERS - CUSTOMER KHATA REPORT"])
    ws.append([f"Customer Name: {customer.name}"])
    ws.append([f"Phone: {customer.phone}"])
    ws.append([f"Net Udhar Balance: Rs. {customer.current_balance}"])
    ws.append([])  # Blank row

    # Table Headers
    headers = [
        "Date & Time",
        "Fertilizer Type",
        "Bags Quantity",
        "Total Amount (Rs.)",
        "Payment Mode",
    ]
    ws.append(headers)

    # Apply styling to table header
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=6, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Table Data
    for txn in transactions:
        row = [
            txn.created_at.strftime("%Y-%m-%d %H:%M"),
            txn.fertilizer_type,
            txn.bags_quantity,
            txn.total_amount,
            txn.payment_status,
        ]
        ws.append(row)

    # Auto-adjust Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to BytesIO stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output